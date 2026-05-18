#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pomi_validate.py — Audits POMI coding against ground-truth production extract.

Loads BQ_Extract_FINAL_*.xlsx, re-codes every priced item using the rewritten
pomi_rules.py + pomi_coder_engine.py, and compares the new NRM L3 code to
what was in the production file.

Usage:
    python3 pomi_validate.py BQ_Extract_FINAL_20260224_192006.xlsx
    python3 pomi_validate.py BQ_Extract_FINAL_*.xlsx --limit 10000  (faster)
    python3 pomi_validate.py BQ_Extract_FINAL_*.xlsx --pomi "POMI_CODING_FINAL 1 (1).xlsx"

Output:
  • Console summary: agreement %, top mismatches, regressions, spot-checks
  • pomi_validation_report.xlsx with sheets:
      SUMMARY  — counts by NRM code, agreement rate
      DIFFS    — every row where new NRM ≠ old NRM (sampled if very many)
      SAMPLES  — 20 random rows from each major NRM section
"""
from __future__ import annotations

import argparse
import os
import random
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Project modules
import pomi_coder_app as app
from pomi_coder_engine import BQCodingEngine

DEFAULT_POMI = 'POMI_CODING_FINAL 1 (1).xlsx'

SPOT_CHECKS = [
    # (description fragment, expected NRM L2 prefix)
    ('pile cap',               '1.1'),
    ('raft foundation',        '1.1'),
    ('suspended slab',         '2.2'),
    ('internal door',          '2.8'),
    ('external door',          '2.6'),
    ('curtain wall',           '2.6'),
    ('porcelain tile; to floor', '3.2'),
    ('porcelain tile; to wall',  '3.1'),
    ('gypsum ceiling',         '3.3'),
    ('wash basin',             '5.1'),
    ('water closet',           '5.1'),
    ('sprinkler',              '5.11'),
    ('smoke detector',         '5.11'),
    ('cctv',                   '5.12'),
    ('rj45',                   '5.12'),
    ('fcu',                    '5.6'),
    ('boiler',                 '5.5'),
    ('exhaust duct',           '5.7'),
    ('passenger elevator',     '5.10'),
    ('escalator',              '5.10'),
]


def normalize_l3(code) -> str:
    """Normalize NRM code to dotted form with no leading zeros."""
    if code is None:
        return ''
    s = str(code).strip()
    if not s:
        return ''
    # Detect placeholders
    if s in ('P', 'PS', 'DW', 'U', '0'):
        return s
    parts = []
    for p in s.split('.'):
        if p.isdigit():
            parts.append(str(int(p)))
        else:
            parts.append(p)
    return '.'.join(parts)


def l2_prefix(code: str) -> str:
    """Return 'L1.L2' from a code like '5.8.2' → '5.8'.  Or '5' → '5'."""
    if not code:
        return ''
    parts = code.split('.')
    if len(parts) >= 2:
        return f'{parts[0]}.{parts[1]}'
    return parts[0]


# Old codes that are KNOWN WRONG. A move away from these counts as IMPROVEMENT.
KNOWN_WRONG_OLD = {
    'P', 'PS', 'DW', 'U', '0',  # placeholders
}
# Old L2 prefixes that were systematically wrong for many domains
WRONG_L2_PATTERNS = {
    '8.1': ('substructure-piling', '1.1'),   # piling tagged as external works
    '8.4': ('foundation', '1.1'),
    '3.1': ('floor-finishes', '3.2'),
}


def classify_diff(old_code: str, new_code: str, old_l2: str, new_l2: str,
                  description: str = '') -> str:
    """Classify (old → new) NRM transition as IMPROVEMENT / REGRESSION / CHANGE."""
    if old_code in KNOWN_WRONG_OLD:
        return 'IMPROVEMENT'
    # If new is a placeholder but old was real → regression
    if new_code in KNOWN_WRONG_OLD and old_code not in KNOWN_WRONG_OLD:
        return 'REGRESSION'
    # If old was a meaningful context code (7.1 demolition, 8.x external,
    # 9.x prelims) and new disagrees on a description that supports the old
    # code, that's a REGRESSION.
    dl = description.lower()
    if old_l2 == '7.1' and any(kw in dl for kw in
            ('demolish', 'demolition', 'cutting of existing', 'removal of existing',
             'alteration', 'strengthening', 'shoring', 'propping')):
        # "Allow for demolition X" → 9.1 (prelim contingency) is acceptable
        if new_l2 == '9.1' and ('allow for' in dl or 'allow demolition' in dl):
            return 'CHANGE'
        if new_l2 != '7.1':
            return 'REGRESSION'
    if old_l2 == '9.1' and any(kw in dl for kw in
            ('the contractor shall', 'allow for', 'provisional', 'dayworks',
             'performance security', 'performance bond', 'insurance')):
        if new_l2 != '9.1':
            return 'REGRESSION'
    return 'CHANGE'


def run_spot_checks(engine, nrm_master) -> list[tuple[str, str, str, bool]]:
    """Run each SPOT_CHECK through the engine; return [(desc, expected, got, pass)]."""
    out = []
    descs = [d for d, _ in SPOT_CHECKS]
    results = engine.code_items(descs, use_ai=False, batch_size=1,
                                contexts=[''] * len(descs),
                                batch_context='spot-check')
    for (desc, expected), r in zip(SPOT_CHECKS, results):
        app.hydrate_nrm_names(r or {}, nrm_master)
        got = ''
        if r and r.get('nrm_code'):
            got = normalize_l3(r['nrm_code'])
        ok = got.startswith(expected) if expected else False
        out.append((desc, expected, got, ok))
    return out


def load_production_rows(xlsx_path: str, limit: int | None = None):
    """Stream the production extract; yield dicts with the fields we need."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['STAGE 1']
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:        # rows 0-3 are header/title
            continue
        if not row or row[1] is None:
            continue
        n += 1
        yield {
            'item_ref':     row[0],
            'description':  str(row[1]) if row[1] is not None else '',
            'qty':          row[2],
            'unit':         row[3],
            'rate':         row[4],
            'amount':       row[5],
            'project_id':   row[14],
            'project_name': row[15],
            'source_file':  row[16],
            'source_sheet': row[17],
            'old_l1':       row[18],
            'old_l1_name':  row[19],
            'old_l2':       row[20],
            'old_l2_name':  row[21],
            'old_l3':       row[22],
            'old_l3_name':  row[23],
        }
        if limit is not None and n >= limit:
            break


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('extract', help='BQ_Extract_FINAL_*.xlsx production extract')
    ap.add_argument('--pomi',  default=DEFAULT_POMI,
                    help=f'POMI_CODING_FINAL workbook (default: {DEFAULT_POMI!r})')
    ap.add_argument('--limit', type=int, default=None,
                    help='Process only the first N data rows (smoke test)')
    ap.add_argument('--out',   default='pomi_validation_report.xlsx')
    ap.add_argument('--max-diffs', type=int, default=5000,
                    help='Cap DIFFS sheet at this many rows (default 5000)')
    args = ap.parse_args()

    if not os.path.exists(args.extract):
        print(f'  ✗  Extract not found: {args.extract}'); sys.exit(1)
    if not os.path.exists(args.pomi):
        print(f'  ✗  POMI workbook not found: {args.pomi}'); sys.exit(1)

    print(f'═══ POMI Validation ═══')
    print(f'  Extract:  {args.extract}')
    print(f'  POMI:     {args.pomi}')
    if args.limit:
        print(f'  Limit:    {args.limit} rows')

    engine = BQCodingEngine(pomi_path=args.pomi, api_key='')
    nrm_master = app._load_nrm_master(args.pomi)
    print(f'  NRM master: {len(nrm_master)} entries')

    # ── Spot checks ─────────────────────────────────────────────────────────
    print(f'\n── Spot checks ──')
    spot = run_spot_checks(engine, nrm_master)
    spot_pass = sum(1 for _, _, _, ok in spot if ok)
    for desc, exp, got, ok in spot:
        flag = '✓' if ok else '✗'
        print(f'  {flag} {desc:<30}  expected={exp:<6}  got={got!r}')
    print(f'  Total: {spot_pass}/{len(spot)} passed')

    # ── Process production extract ──────────────────────────────────────────
    print(f'\n── Streaming production extract ──')
    rows = []
    BATCH = 500
    descriptions: list[str] = []
    batch_rows: list[dict] = []

    total_rows  = 0
    noise_skip  = 0
    coded_ok    = 0
    placeholder_remap = Counter()
    new_nrm_dist = Counter()
    old_nrm_dist = Counter()
    diffs        = []
    agreements   = 0
    disagreements = 0
    stage_counts = Counter()
    improvements = 0
    regressions  = 0
    changes      = 0
    samples_by_l2: dict[str, list[dict]] = defaultdict(list)

    def flush_batch():
        nonlocal coded_ok, agreements, disagreements, improvements, regressions, changes
        if not batch_rows:
            return
        results = engine.code_items(
            descriptions, use_ai=False, batch_size=BATCH,
            contexts=[''] * len(descriptions),
            batch_context='validation')
        for row, r in zip(batch_rows, results):
            if r:
                app.remap_placeholder_nrm(r)
                if app._is_prose_clause(row['description'], row.get('qty'), row.get('unit')):
                    r['nrm_code'] = '9.01'
                    r['nrm_desc'] = 'Preliminaries / General Requirements'
                app.hydrate_nrm_names(r, nrm_master)
                stage_counts[r.get('stage', '?')] += 1
                if r.get('pomi_code'):
                    coded_ok += 1

            new_code = normalize_l3(r['nrm_code']) if r and r.get('nrm_code') else ''
            old_code = normalize_l3(row['old_l3'])
            new_l2 = l2_prefix(new_code)
            old_l2 = l2_prefix(old_code)

            if old_code in app.PLACEHOLDER_NRM_MAP:
                placeholder_remap[old_code] += 1

            new_nrm_dist[new_l2 or 'unmatched'] += 1
            old_nrm_dist[old_l2 or 'unmatched'] += 1

            if new_l2 and old_l2 and new_l2 == old_l2:
                agreements += 1
            elif new_l2 and old_l2:
                disagreements += 1
                verdict = classify_diff(old_code, new_code, old_l2, new_l2,
                                        description=row['description'])
                if verdict == 'IMPROVEMENT': improvements += 1
                elif verdict == 'REGRESSION': regressions += 1
                else: changes += 1
                if len(diffs) < args.max_diffs:
                    diffs.append({
                        'source_file': row['source_file'],
                        'project':     row['project_name'],
                        'desc':        row['description'][:200],
                        'qty':         row['qty'],
                        'unit':        row['unit'],
                        'old_l3':      old_code,
                        'new_l3':      new_code,
                        'old_l2_name': row['old_l2_name'] or '',
                        'new_l2_name': r.get('nrm_l2_name', '') if r else '',
                        'new_l3_name': r.get('nrm_l3_name', '') if r else '',
                        'stage':       r.get('stage', '') if r else '',
                        'confidence':  r.get('confidence', '') if r else '',
                        'verdict':     verdict,
                    })
            # Collect samples by L2 group
            if new_l2 and len(samples_by_l2[new_l2]) < 20 and random.random() < 0.02:
                samples_by_l2[new_l2].append({
                    'desc':    row['description'][:200],
                    'old_l3':  old_code,
                    'new_l3':  new_code,
                    'new_l2_name': r.get('nrm_l2_name', '') if r else '',
                    'confidence':  r.get('confidence', '') if r else '',
                    'stage':       r.get('stage', '') if r else '',
                })

        batch_rows.clear()
        descriptions.clear()

    for row in load_production_rows(args.extract, limit=args.limit):
        total_rows += 1
        # Skip noise rows
        if app._is_noise_row(row['description'], row.get('qty'), row.get('amount')):
            noise_skip += 1
            continue
        batch_rows.append(row)
        descriptions.append(row['description'])
        if len(batch_rows) >= BATCH:
            flush_batch()
        if total_rows % 5000 == 0:
            print(f'  …processed {total_rows:>8,} rows | coded={coded_ok:>7,} | '
                  f'agreement={agreements:>7,} | diffs={disagreements:>7,}')
    flush_batch()

    print(f'\n── Results ──')
    print(f'  Total data rows:         {total_rows:,}')
    print(f'  Filtered as noise:       {noise_skip:,}')
    print(f'  Coded successfully:      {coded_ok:,}')
    print(f'  Stages:                  {dict(stage_counts)}')
    coded_compared = agreements + disagreements
    pct = (agreements / coded_compared * 100) if coded_compared else 0
    print(f'  Agreement @ L2:          {agreements:,} / {coded_compared:,}  ({pct:.1f}%)')
    print(f'  Diffs classified:        IMP={improvements:,}  REG={regressions:,}  CHG={changes:,}')
    print(f'  Placeholders remapped:   {dict(placeholder_remap)}')
    print()
    print(f'  Top NEW NRM L2 distribution:')
    for c, n in new_nrm_dist.most_common(10):
        print(f'    {c:>8}  {n:>7,}')
    print(f'  Spot checks:             {spot_pass}/{len(spot)}')

    # ── Write report workbook ───────────────────────────────────────────────
    print(f'\n── Writing report: {args.out} ──')
    wb = Workbook()
    wb.remove(wb.active)

    def fnt(bold=False, size=10, col='FF000000'):
        return Font(name='Arial', bold=bold, size=size, color=col)
    def fill(h): return PatternFill('solid', fgColor=h)
    def aln(h='left'): return Alignment(horizontal=h, vertical='center', wrap_text=False)
    side = Side(style='thin', color='FFD0D0D0')
    def bdr(): return Border(left=side, right=side, top=side, bottom=side)

    # SUMMARY
    ws = wb.create_sheet('SUMMARY')
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18
    summary_rows = [
        ('Extract',                       os.path.basename(args.extract)),
        ('POMI workbook',                 os.path.basename(args.pomi)),
        ('Rules loaded',                  len(engine._all_rules)),
        ('NRM master entries',            len(nrm_master)),
        ('Total data rows',               total_rows),
        ('Filtered as noise',             noise_skip),
        ('Coded successfully',            coded_ok),
        ('Agreement @ L2',                f'{agreements:,} / {coded_compared:,}'),
        ('Agreement %',                   f'{pct:.1f}%'),
        ('Improvements',                  improvements),
        ('Regressions',                   regressions),
        ('Other changes',                 changes),
        ('Stage: rule',                   stage_counts.get('rule', 0)),
        ('Stage: fuzzy',                  stage_counts.get('fuzzy', 0)),
        ('Stage: ai',                     stage_counts.get('ai', 0)),
        ('Stage: unmatched',              stage_counts.get('unmatched', 0)),
        ('Spot-checks passed',            f'{spot_pass}/{len(spot)}'),
        ('Placeholder P → 9.01',          placeholder_remap.get('P', 0)),
        ('Placeholder PS → 9.01',         placeholder_remap.get('PS', 0)),
        ('Placeholder DW → 9.01',         placeholder_remap.get('DW', 0)),
        ('Placeholder 0 → 7.01',          placeholder_remap.get('0', 0)),
    ]
    for i, (k, v) in enumerate(summary_rows, 1):
        ws.cell(i, 1, k).font = fnt(bold=True)
        ws.cell(i, 1).fill = fill('FFEEF5FB')
        ws.cell(i, 2, v).font = fnt()
    # NEW NRM distribution
    base = len(summary_rows) + 2
    ws.cell(base, 1, 'NEW NRM L2 Distribution').font = fnt(bold=True, col='FF1A5276')
    ws.cell(base, 1).fill = fill('FFD6EAF8')
    ws.cell(base+1, 1, 'L2 Code').font = fnt(bold=True)
    ws.cell(base+1, 2, 'Count').font = fnt(bold=True)
    for j, (code, count) in enumerate(new_nrm_dist.most_common(40), base + 2):
        ws.cell(j, 1, code)
        ws.cell(j, 2, count)

    # Spot-check sheet
    ws = wb.create_sheet('SPOT_CHECKS')
    for col, w in [('A', 36), ('B', 12), ('C', 14), ('D', 8)]:
        ws.column_dimensions[col].width = w
    hdr = ['Description', 'Expected L2', 'Got L3', 'Pass']
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(1, ci, h)
        c.font = fnt(bold=True, col='FFFFFFFF'); c.fill = fill('FF1A5276')
        c.border = bdr(); c.alignment = aln('center')
    for ri, (desc, exp, got, ok) in enumerate(spot, 2):
        ws.cell(ri, 1, desc).border = bdr()
        ws.cell(ri, 2, exp).border = bdr()
        ws.cell(ri, 3, got).border = bdr()
        cflag = ws.cell(ri, 4, '✓' if ok else '✗')
        cflag.border = bdr()
        cflag.fill = fill('FFD5F5E3' if ok else 'FFFDE8D8')
        cflag.font = fnt(bold=True, col='FF1E8449' if ok else 'FFB03A2E')
        cflag.alignment = aln('center')

    # DIFFS
    ws = wb.create_sheet('DIFFS')
    diffs_cols = [
        ('A', 'Source File',  32),
        ('B', 'Project',      18),
        ('C', 'Description',  60),
        ('D', 'Qty',           8),
        ('E', 'Unit',          8),
        ('F', 'Old L3',        9),
        ('G', 'New L3',        9),
        ('H', 'Old L2 Name',  26),
        ('I', 'New L2 Name',  26),
        ('J', 'New L3 Name',  32),
        ('K', 'Stage',         8),
        ('L', 'Conf',          7),
        ('M', 'Verdict',      12),
    ]
    for col, hdr_, w in diffs_cols:
        ws.column_dimensions[col].width = w
    for ci, (_, hdr_, _) in enumerate(diffs_cols, 1):
        c = ws.cell(1, ci, hdr_); c.font = fnt(bold=True, col='FFFFFFFF')
        c.fill = fill('FFB03A2E'); c.border = bdr(); c.alignment = aln('center')
    for ri, d in enumerate(diffs, 2):
        vals = [d['source_file'], d['project'], d['desc'], d['qty'], d['unit'],
                d['old_l3'], d['new_l3'], d['old_l2_name'], d['new_l2_name'],
                d['new_l3_name'], d['stage'], d['confidence'], d['verdict']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v); c.border = bdr(); c.font = fnt(size=9)
        c = ws.cell(ri, 13)
        if d['verdict'] == 'IMPROVEMENT':
            c.fill = fill('FFD5F5E3'); c.font = fnt(bold=True, size=9, col='FF1E8449')
        elif d['verdict'] == 'REGRESSION':
            c.fill = fill('FFFDE8D8'); c.font = fnt(bold=True, size=9, col='FFB03A2E')

    # SAMPLES
    ws = wb.create_sheet('SAMPLES')
    samples_cols = [
        ('A', 'L2 Group',    10),
        ('B', 'L2 Name',     30),
        ('C', 'Description', 60),
        ('D', 'Old L3',       9),
        ('E', 'New L3',       9),
        ('F', 'Stage',        8),
        ('G', 'Conf',         7),
    ]
    for col, hdr_, w in samples_cols:
        ws.column_dimensions[col].width = w
    for ci, (_, hdr_, _) in enumerate(samples_cols, 1):
        c = ws.cell(1, ci, hdr_); c.font = fnt(bold=True, col='FFFFFFFF')
        c.fill = fill('FF1A5276'); c.border = bdr(); c.alignment = aln('center')
    ri = 2
    for l2, lst in sorted(samples_by_l2.items()):
        for s in lst[:20]:
            vals = [l2, s['new_l2_name'], s['desc'], s['old_l3'], s['new_l3'],
                    s['stage'], s['confidence']]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(ri, ci, v); c.border = bdr(); c.font = fnt(size=9)
            ri += 1

    wb.save(args.out)
    print(f'  Report saved: {args.out}')
    print('\n═══ Validation complete ═══')


if __name__ == '__main__':
    main()
