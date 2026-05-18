#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dwg_to_boq.py — Extract a BoQ directly FROM a DWG/DXF drawing (or folder).

Reuses the layer→NRM mapping and the extractor from
python/extract_dwg_elements.py. Produces an Excel workbook with:

  • BOQ ITEMS    — one row per NRM L2 (count or qty + unit + confidence)
  • ICMS SUMMARY — rolled up by ICMS L3 group
  • DETAIL       — every individual entity with block name + position + source

Usage:
    python3 dwg_to_boq.py PATH.dwg
    python3 dwg_to_boq.py PATH.dwg --out my_takeoff.xlsx
    python3 dwg_to_boq.py /folder/of/drawings/ --out project_takeoff.xlsx
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).parent / "python"))
from extract_dwg_elements import extract_dwg_elements


ROOT = Path(__file__).parent
DEFAULT_LAYER_MAP = ROOT / "dictionaries" / "layer_to_nrm.default.json"
DEFAULT_ICMS_MAP = ROOT / "dictionaries" / "nrm_to_icms.default.json"


# ── style helpers ─────────────────────────────────────────────────────────
NAVY = "FF1B2A4A"


def sol(h):
    return PatternFill("solid", fgColor=h)


def fnt(bold=False, size=10, col="FF000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=col, italic=italic)


def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def bdr(col="FFD0D0D0"):
    s = Side(style="thin", color=col)
    return Border(left=s, right=s, top=s, bottom=s)


# ── core ──────────────────────────────────────────────────────────────────
def collect_drawing_paths(arg):
    """Return sorted list of .dwg/.dxf absolute paths from a file or dir."""
    p = os.path.abspath(arg)
    if os.path.isfile(p):
        return [p]
    if not os.path.isdir(p):
        raise FileNotFoundError(p)
    out = []
    for root, _, files in os.walk(p):
        for fn in files:
            if fn.lower().endswith((".dwg", ".dxf")) and not fn.startswith("."):
                out.append(os.path.join(root, fn))
    return sorted(out)


def aggregate(drawings, layer_map_path):
    """Run extractor over each drawing; combine into per-NRM totals.
    Returns (combined, per_drawing_extracts)."""
    combined = {}   # nrm_l2 → bucket dict
    per_drawing = []

    for path in drawings:
        try:
            ex = extract_dwg_elements(path, layer_map_path, verbose=False)
        except Exception as e:
            print(f"  ⚠  extract failed: {os.path.basename(path)} — {e}", file=sys.stderr)
            continue
        per_drawing.append({"source": path, "extract": ex})

        for nrm, b in ex.by_nrm.items():
            c = combined.setdefault(nrm, {
                "nrm_l2": nrm,
                "category": b.category,
                "unit": b.unit,
                "confidence": b.confidence,
                "count": 0,
                "qty_total": 0.0,
                "sources": [],
                "items": [],
            })
            c["count"] += b.count
            c["qty_total"] += b.qty_total
            c["sources"].append(os.path.basename(path))
            c["items"].extend([
                {**it, "_source": os.path.basename(path)} for it in b.items
            ])
    return combined, per_drawing


# NRM L1 names so we can hydrate L1 / L2 on every row
NRM_L1_NAMES = {
    "1": "Substructure",
    "2": "Superstructure",
    "3": "Internal Finishes",
    "4": "Fittings, Furnishings and Equipment",
    "5": "Services",
    "6": "Prefabricated Buildings and Building Units",
    "7": "Work to Existing Buildings",
    "8": "External Works",
    "9": "Preliminaries / General Requirements",
}


def load_nrm_master(pomi_workbook_path):
    """Best-effort NRM L2 → name lookup from the POMI master workbook."""
    if not pomi_workbook_path or not os.path.exists(pomi_workbook_path):
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(pomi_workbook_path, data_only=True, read_only=True)
        if "NRM" not in wb.sheetnames:
            return {}
        out = {}
        ws = wb["NRM"]
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 17:
                continue
            code, name = row[15], row[16]
            if code and name:
                parts = [str(int(p)) if str(p).isdigit() else str(p)
                         for p in str(code).split(".")]
                norm = ".".join(parts)
                out[norm] = str(name).strip()
        return out
    except Exception:
        return {}


def load_icms_mapping(icms_path=DEFAULT_ICMS_MAP):
    try:
        with open(icms_path) as f:
            return json.load(f)
    except Exception:
        return None


# ── writers ───────────────────────────────────────────────────────────────
def write_boq_sheet(wb, combined, source_label, nrm_names):
    ws = wb.create_sheet("BOQ ITEMS")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 95

    cols = [
        ("A", "Item",          7),
        ("B", "NRM L1",        7),
        ("C", "NRM L1 Name",  28),
        ("D", "NRM L2",        8),
        ("E", "NRM L2 Name",  32),
        ("F", "Category",     30),
        ("G", "Description",  46),
        ("H", "Count",        10),
        ("I", "Qty",          12),
        ("J", "Unit",          7),
        ("K", "Confidence",   12),
        ("L", "Sources",      40),
    ]
    for cl, _, w in cols:
        ws.column_dimensions[cl].width = w

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = f"BoQ (extracted from drawings)  ·  {source_label}"
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFFFF")
    c.fill = sol(NAVY)
    c.alignment = aln("center", "center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    c = ws["A2"]
    c.value = ("Layer-based take-off — every priced item below corresponds to a NRM L2 group "
               "populated from CAD layer codes.  See DETAIL sheet for entity-level breakdown.")
    c.font = Font(name="Arial", size=9, color="FF566573", italic=True)
    c.fill = sol("FFF8F9FA")
    c.alignment = aln("left", "center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for ci, (_, hdr, _) in enumerate(cols, 1):
        c = ws.cell(4, ci)
        c.value = hdr
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFFFF")
        c.fill = sol(NAVY)
        c.alignment = aln("center", "center", True)
        c.border = bdr()
    ws.row_dimensions[4].height = 24

    CONF_COLOR = {
        "HIGH":   ("FFD5F5E3", "FF1E8449"),
        "MEDIUM": ("FFFFFBE6", "FF7D6608"),
        "LOW":    ("FFFDE8D8", "FFB03A2E"),
    }

    out_r = 5
    item_idx = 1
    for nrm in sorted(combined.keys()):
        b = combined[nrm]
        l1 = nrm.split(".")[0]
        l1_name = NRM_L1_NAMES.get(l1, "")
        l2_name = nrm_names.get(nrm, b["category"])
        sources = ", ".join(sorted(set(b["sources"])))[:80]
        rf = sol("FFFFFFFF") if out_r % 2 == 0 else sol("FFF8F9FA")
        cfill, cfont = CONF_COLOR.get(b["confidence"], ("FFEAEAEA", "FF666666"))

        # Description = "Category from layer; count INSERTs / qty m..."
        if b["unit"] == "Nr":
            desc = f"{b['category']} (count from layer blocks)"
            qty_display = b["count"]
        else:
            desc = f"{b['category']} (length/area from layer geometry)"
            qty_display = round(b["qty_total"], 2)

        vals = [
            item_idx, l1, l1_name, nrm, l2_name, b["category"], desc,
            b["count"] if b["count"] else "",
            qty_display if qty_display else "",
            b["unit"], b["confidence"], sources,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(out_r, ci)
            c.value = v
            c.border = bdr()
            c.fill = rf
            c.font = fnt(size=9)
            if ci in (1, 2, 4, 8, 9, 10, 11):
                c.alignment = aln("center", "center")
            elif ci == 9 and isinstance(v, (int, float)):
                c.number_format = "#,##0.00"
            else:
                c.alignment = aln("left", "center", True)
            if ci == 11:  # confidence
                c.fill = sol(cfill)
                c.font = fnt(bold=True, size=9, col=cfont)
        ws.row_dimensions[out_r].height = 20
        out_r += 1
        item_idx += 1

    out_r += 1
    ws.merge_cells(f"A{out_r}:G{out_r}")
    c = ws.cell(out_r, 1)
    c.value = f"TOTAL  ·  {item_idx-1} NRM L2 groups populated"
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("right", "center")
    total_count = sum(b["count"] for b in combined.values())
    total_qty = sum(b["qty_total"] for b in combined.values() if b["unit"] != "Nr")
    c = ws.cell(out_r, 8)
    c.value = total_count
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("center", "center")
    c = ws.cell(out_r, 9)
    c.value = round(total_qty, 2) if total_qty else ""
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("center", "center")
    ws.row_dimensions[out_r].height = 26


def write_icms_sheet(wb, combined, icms_data, source_label):
    if icms_data is None:
        return
    ws = wb.create_sheet("ICMS SUMMARY")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 95

    cols = [("A", "ICMS Group", 14), ("B", "ICMS Code", 11),
            ("C", "Group Name", 46), ("D", "Items", 8),
            ("E", "Total Count", 14), ("F", "% of Items", 12)]
    for cl, _, w in cols:
        ws.column_dimensions[cl].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"ICMS SUMMARY (from drawings)  ·  {source_label}"
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFFFF")
    c.fill = sol(NAVY)
    c.alignment = aln("center", "center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "International Construction Measurement Standards (3rd ed., 2021)"
    c.font = Font(name="Arial", size=9, color="FF566573", italic=True)
    c.fill = sol("FFF8F9FA")
    c.alignment = aln("left", "center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    for ci, (_, hdr, _) in enumerate(cols, 1):
        c = ws.cell(4, ci)
        c.value = hdr
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFFFF")
        c.fill = sol(NAVY)
        c.alignment = aln("center", "center", True)
        c.border = bdr()
    ws.row_dimensions[4].height = 24

    mapping = icms_data.get("mapping", {})
    all_groups = icms_data.get("all_icms_groups_buildings", {})

    by_group = defaultdict(lambda: {"items": 0, "count": 0})
    for nrm, b in combined.items():
        entry = mapping.get(nrm, {})
        gc = entry.get("icms_l3", "")
        gn = entry.get("icms_group_name", "Unclassified")
        if not gc:
            gc, gn = "—", "Unclassified (no ICMS mapping)"
        by_group[(gc, gn)]["items"] += 1
        by_group[(gc, gn)]["count"] += b["count"]

    grand_count = sum(g["count"] for g in by_group.values()) or 1

    out_r = 5
    GROUP_COLORS = {
        "01": "FF7D6608", "02": "FF1A5276", "03": "FF1B2A4A", "04": "FF1B7A4E",
        "05": "FF117A65", "06": "FF1A6B75", "07": "FF6E2F0A", "08": "FF4A235A",
        "09": "FF2C3E50", "10": "FF2C3E50", "11": "FF566573", "12": "FF943126",
        "13": "FF7B241C",
    }
    for gc in sorted(all_groups.keys()):
        gn = all_groups[gc]
        match = next((d for (k, _), d in by_group.items() if k == gc), None)
        items_n = match["items"] if match else 0
        count_n = match["count"] if match else 0
        pct = count_n / grand_count
        rf = sol("FFFFFFFF") if out_r % 2 == 0 else sol("FFF8F9FA")
        color = GROUP_COLORS.get(gc, "FF566573")
        is_filled = items_n > 0
        vals = ["CC", gc, gn,
                items_n if items_n else "",
                count_n if count_n else "",
                pct if is_filled else ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(out_r, ci)
            c.value = v
            c.border = bdr()
            c.fill = rf
            c.font = fnt(size=9)
            if ci in (1, 2, 4, 5):
                c.font = fnt(bold=is_filled, size=10, col=color if is_filled else "FF999999")
                c.alignment = aln("center", "center")
            elif ci == 3:
                c.font = fnt(bold=is_filled, size=10,
                             col="FF1B2A4A" if is_filled else "FF999999")
                c.alignment = aln("left", "center")
            elif ci == 6:
                c.font = fnt(bold=is_filled, size=10,
                             col=color if is_filled else "FF999999")
                c.alignment = aln("right", "center")
                if isinstance(v, (int, float)):
                    c.number_format = "0.0%"
        ws.row_dimensions[out_r].height = 20
        out_r += 1

    out_r += 1
    ws.merge_cells(f"A{out_r}:D{out_r}")
    c = ws.cell(out_r, 1)
    c.value = "GRAND TOTAL  ·  CC items in drawing"
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("right", "center")
    c = ws.cell(out_r, 5)
    c.value = grand_count
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("center", "center")
    c = ws.cell(out_r, 6)
    c.value = 1.0
    c.number_format = "0.0%"
    c.font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    c.fill = sol("FFCA6F1E")
    c.alignment = aln("right", "center")
    ws.row_dimensions[out_r].height = 26


def write_detail_sheet(wb, combined):
    ws = wb.create_sheet("DETAIL")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 95

    cols = [("A", "NRM L2", 8), ("B", "Category", 30), ("C", "Entity Type", 14),
            ("D", "Block Name", 30), ("E", "Layer", 60),
            ("F", "X", 14), ("G", "Y", 14), ("H", "Qty", 12),
            ("I", "Source", 38)]
    for cl, _, w in cols:
        ws.column_dimensions[cl].width = w
    for ci, (_, hdr, _) in enumerate(cols, 1):
        c = ws.cell(1, ci)
        c.value = hdr
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFFFF")
        c.fill = sol(NAVY)
        c.alignment = aln("center", "center", True)
        c.border = bdr()
    ws.row_dimensions[1].height = 22

    out_r = 2
    for nrm in sorted(combined.keys()):
        b = combined[nrm]
        for it in b["items"]:
            pt = it.get("insertion_point") or [None, None]
            vals = [nrm, b["category"], it.get("type", ""),
                    it.get("block_name", ""), it.get("layer", ""),
                    pt[0] if pt and pt[0] is not None else "",
                    pt[1] if pt and pt[1] is not None else "",
                    it.get("qty", "") if it.get("type") == "LWPOLYLINE" else "",
                    it.get("_source", "")]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(out_r, ci)
                c.value = v
                c.border = bdr()
                c.font = fnt(size=8)
                c.alignment = aln("center" if ci in (1, 3, 6, 7, 8) else "left", "center", True)
                if ci in (6, 7) and isinstance(v, (int, float)):
                    c.number_format = "#,##0.00"
            out_r += 1


# ── main ─────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="Extract a BoQ directly from a DWG/DXF drawing or folder.")
    ap.add_argument("drawing", help="Path to .dwg/.dxf OR a folder of drawings")
    ap.add_argument("--out", default="", help="Output workbook (default: <input>_TAKEOFF.xlsx)")
    ap.add_argument("--layers", default=str(DEFAULT_LAYER_MAP),
                    help=f"Layer→NRM mapping (default: {DEFAULT_LAYER_MAP.name})")
    ap.add_argument("--icms", default=str(DEFAULT_ICMS_MAP),
                    help=f"NRM→ICMS mapping (default: {DEFAULT_ICMS_MAP.name})")
    ap.add_argument("--pomi", default="",
                    help="Optional POMI_CODING_FINAL.xlsx for NRM L2 names")
    args = ap.parse_args()

    drawings = collect_drawing_paths(args.drawing)
    if not drawings:
        print(f"  ✗  no drawings found in {args.drawing}", file=sys.stderr)
        sys.exit(1)

    label = os.path.basename(os.path.abspath(args.drawing))
    out_path = args.out or (os.path.splitext(os.path.abspath(args.drawing))[0] + "_TAKEOFF.xlsx")

    print(f"═══ DWG → BoQ ═══")
    print(f"  Input:    {label}")
    print(f"  Drawings: {len(drawings)}")
    for p in drawings:
        print(f"    • {os.path.basename(p)}")
    print(f"  Output:   {out_path}")
    print()

    combined, per_drawing = aggregate(drawings, args.layers)
    print(f"  NRM L2 groups populated: {len(combined)}")
    for nrm in sorted(combined.keys()):
        b = combined[nrm]
        qty_str = f"  qty={b['qty_total']:>8.2f} {b['unit']}" if b["unit"] != "Nr" else ""
        print(f"    {nrm:<5}  {b['category']:<32}  count={b['count']:>4}  {b['confidence']:<7}{qty_str}")

    nrm_names = load_nrm_master(args.pomi)
    icms_data = load_icms_mapping(args.icms)

    wb = Workbook()
    wb.remove(wb.active)
    write_boq_sheet(wb, combined, label, nrm_names)
    write_icms_sheet(wb, combined, icms_data, label)
    write_detail_sheet(wb, combined)
    wb.save(out_path)

    print(f"\n  ✅  Saved: {out_path}")


if __name__ == "__main__":
    main()
