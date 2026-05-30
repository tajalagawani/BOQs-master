#!/usr/bin/env python3
"""strip_pomi_columns.py — Strip POMI-added columns from an
already-coded BoQ to recover the original bill structure, so the
pipeline can re-process it fresh.

Detects the structural shape:
  * Header row has 'Row' / 'REF' / 'Description' / 'Qty' / 'Unit' /
    'Rate' / 'Amount' (in that order).
  * Drops column 1 ('Row' — internal counter).
  * Keeps cols 2..7 (REF, Description, Qty, Unit, Rate, Amount).
  * Strips the row-prefix marker ('  ▸  ') the coder added on info
    rows so the new run sees the same description shape as the
    original BoQ.

Pure structural; no sheet-name or project-specific lists.
"""
from __future__ import annotations
import argparse, os, re
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# Marker the coder adds in front of context/info-bar rows.
_PREFIX_RE = re.compile(r"^\s*▸\s*")


def find_header_row(ws, max_scan: int = 8) -> int | None:
    """Locate the row that holds the column headers 'Description'/'Qty'/'Unit'."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = [str(ws.cell(row=r, column=c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)]
        if "description" in vals and "qty" in vals and "unit" in vals:
            return r
    return None


def column_indexes(ws, hdr_row: int) -> dict:
    """Return {logical_name: 1-based col index} for the original BoQ cols."""
    names = {"row": None, "ref": None, "description": None,
             "qty": None, "unit": None, "rate": None, "amount": None}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=hdr_row, column=c).value or "").strip().lower()
        if v in names and names[v] is None:
            names[v] = c
    return names


def strip_sheet(src_ws, dst_ws) -> int:
    """Copy original-BoQ cells from src into a fresh sheet; returns rows written."""
    hdr = find_header_row(src_ws)
    if hdr is None:
        return 0
    cols = column_indexes(src_ws, hdr)
    # We need at least description+qty+unit to be a usable sheet
    if not (cols["description"] and cols["qty"] and cols["unit"]):
        return 0

    # Output layout (standard pipeline-friendly):
    #   A=Item, B=Description, C=Quantity, D=Unit, E=Rate, F=Amount
    dst_ws.append(["Item", "Description", "Quantity", "Unit", "Rate", "Amount"])
    written = 1
    for r in range(hdr + 1, src_ws.max_row + 1):
        ref  = src_ws.cell(row=r, column=cols["ref"]).value         if cols["ref"]         else None
        desc = src_ws.cell(row=r, column=cols["description"]).value if cols["description"] else None
        qty  = src_ws.cell(row=r, column=cols["qty"]).value         if cols["qty"]         else None
        unit = src_ws.cell(row=r, column=cols["unit"]).value        if cols["unit"]        else None
        rate = src_ws.cell(row=r, column=cols["rate"]).value        if cols["rate"]        else None
        amt  = src_ws.cell(row=r, column=cols["amount"]).value      if cols["amount"]      else None
        # Strip the '  ▸  ' marker the coder added on info rows
        if isinstance(desc, str):
            desc = _PREFIX_RE.sub("", desc).rstrip()
        dst_ws.append([ref, desc, qty, unit, rate, amt])
        written += 1
    return written


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("source", help="Coded BoQ to strip")
    ap.add_argument("--out", default="", help="Output xlsx (default: <name>_RAW.xlsx)")
    args = ap.parse_args()

    src_path = os.path.abspath(args.source)
    if not args.out:
        base, ext = os.path.splitext(src_path)
        args.out = f"{base}_RAW{ext}"

    wb_in = load_workbook(src_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    kept = 0
    skipped = []
    for sname in wb_in.sheetnames:
        src_ws = wb_in[sname]
        # Skip pipeline-added sheets — they don't represent a bill
        if sname in ("MASTER BOQs", "ICMS SUMMARY", "RECONCILIATION",
                     "RATE_ONLY", "UNPRICED", "SUMMARY",
                     "⚠ FLAGGED FOR REVIEW"):
            skipped.append(sname)
            continue
        dst_ws = wb_out.create_sheet(sname[:31])
        n = strip_sheet(src_ws, dst_ws)
        if n <= 1:
            wb_out.remove(dst_ws)
            skipped.append(f"{sname} (no usable rows)")
        else:
            kept += 1
            print(f"  ✓ {sname:<25}  {n:>5} rows")
    if not wb_out.sheetnames:
        raise SystemExit("error: no sheets recovered")

    wb_out.save(args.out)
    print(f"\n  kept    : {kept} sheets")
    print(f"  skipped : {len(skipped)} ({', '.join(skipped[:8])}{'…' if len(skipped) > 8 else ''})")
    print(f"  saved   : {args.out}")


if __name__ == "__main__":
    main()
